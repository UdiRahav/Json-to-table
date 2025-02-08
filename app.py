from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
from io import BytesIO
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert():
    try:
        data = request.get_json()
        print("Received data:", json.dumps(data, indent=2))
        
        if not data or 'insights' not in data:
            print("No valid insights data found")
            return jsonify({'error': 'No valid insights data provided'}), 400
        
        insights = data['insights']
        print(f"Processing {len(insights)} insights")
        processed_insights = []
        
        for i, insight in enumerate(insights):
            print(f"Processing insight {i + 1}")
            processed_facts = {}
            if 'facts' in insight:
                for fact_name, fact_data in insight['facts'].items():
                    print(f"Processing fact: {fact_name}")
                    if isinstance(fact_data, dict) and 'cols' in fact_data and 'rows' in fact_data:
                        processed_facts[fact_name] = {
                            'headers': fact_data['cols'],
                            'rows': fact_data['rows'],
                            'type': fact_data.get('type', ''),
                            'attributesTypes': fact_data.get('attributesTypes', [])
                        }
            
            processed_insight = {
                'id': insight.get('id', ''),
                'insightId': insight.get('insightId', ''),
                'useCaseId': insight.get('useCaseId', ''),
                'segment': insight.get('segment', ''),
                'generatedDate': insight.get('generatedDate', ''),
                'score': insight.get('score', ''),
                'status': insight.get('status', ''),
                'type': insight.get('type', ''),
                'facts': processed_facts
            }
            print(f"Processed insight: {json.dumps(processed_insight, indent=2)}")
            processed_insights.append(processed_insight)

        result = {
            'requestId': data.get('requestId', ''),
            'insights': processed_insights
        }
        print(f"Returning {len(processed_insights)} insights")
        return jsonify(result)
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {str(e)}")
        return jsonify({'error': 'Invalid JSON format'}), 400
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/favicon.ico')
def favicon():
    return app.send_from_directory('static', 'favicon.ico')

@app.route('/download', methods=['POST'])
def download_excel():
    try:
        request_data = request.get_json()
        if not request_data or 'data' not in request_data:
            print("No data provided")
            return jsonify({'error': 'No data provided'}), 400
            
        data = request_data['data']
        filename = request_data.get('filename', 'insights.xlsx')
        
        # Ensure filename has .xlsx extension
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'

        print(f"Creating Excel file: {filename}")
        
        # Create Excel file in memory
        output = BytesIO()
        
        try:
            with pd.ExcelWriter(output, engine='openpyxl', mode='w') as writer:
                insights = data.get('insights', [])
                print(f"Processing {len(insights)} insights")
                
                if not insights:
                    print("No insights found")
                    df = pd.DataFrame({'Message': ['No data available']})
                    df.to_excel(writer, sheet_name='No Data', index=False)
                    return
                
                # Group insights by use case ID
                use_case_insights = {}
                for insight in insights:
                    use_case_id = insight.get('useCaseId', 'Unknown')
                    if use_case_id not in use_case_insights:
                        use_case_insights[use_case_id] = []
                    use_case_insights[use_case_id].append(insight)
                
                # Process each use case
                for use_case_id, case_insights in use_case_insights.items():
                    print(f"Processing use case: {use_case_id}")
                    
                    # Create a list to store all DataFrames and their styles for this use case
                    all_dfs = []
                    
                    # Process each insight in this use case
                    for i, insight in enumerate(case_insights, 1):
                        facts = insight.get('facts', {})
                        
                        # Add insight header
                        header = pd.DataFrame({
                            '': [f'Insight {i} Details']
                        })
                        all_dfs.append((header, 'header'))
                        
                        # Add insight details as a small table
                        insight_details = pd.DataFrame({
                            'Property': [
                                'Insight ID', 'Type', 'Segment', 
                                'Score', 'Status', 'Generated Date'
                            ],
                            'Value': [
                                insight.get('insightId', ''),
                                insight.get('type', ''),
                                insight.get('segment', ''),
                                insight.get('score', ''),
                                insight.get('status', ''),
                                insight.get('generatedDate', '')
                            ]
                        })
                        all_dfs.append((insight_details, 'details'))
                        
                        # Add a separator
                        separator = pd.DataFrame({'': ['']})
                        all_dfs.append((separator, 'separator'))
                        
                        # Process each fact
                        for fact_name, fact_data in facts.items():
                            if isinstance(fact_data, dict) and 'cols' in fact_data and 'rows' in fact_data:
                                print(f"Processing fact: {fact_name}")
                                
                                # Add fact title
                                title = pd.DataFrame({
                                    '': [f'Fact: {fact_name}']
                                })
                                all_dfs.append((title, 'fact_title'))
                                
                                # Create DataFrame from fact data
                                df = pd.DataFrame(fact_data['rows'], columns=fact_data['cols'])
                                
                                # Try to convert date columns and sort
                                date_columns = []
                                for col in df.columns:
                                    if any(date_word in col.lower() for date_word in ['date', 'time', 'day']):
                                        try:
                                            # Convert to datetime and then to string in a consistent format
                                            dates = pd.to_datetime(df[col])
                                            df[col] = dates.dt.strftime('%Y-%m-%d %H:%M:%S')
                                            date_columns.append(col)
                                        except Exception as e:
                                            print(f"Failed to convert column {col} to datetime: {str(e)}")
                                            continue
                                
                                # Sort by the first date column found if any exist
                                if date_columns:
                                    sort_col = date_columns[0]
                                    # Temporarily convert back to datetime for sorting
                                    temp_dates = pd.to_datetime(df[sort_col])
                                    df = df.iloc[temp_dates.argsort()]
                                
                                # Add the fact data
                                all_dfs.append((df, 'fact_data'))
                                
                                # Add a separator
                                all_dfs.append((separator, 'separator'))
                    
                    if all_dfs:
                        # Combine all DataFrames with spaces between them
                        sheet_name = str(use_case_id)[:31]  # Excel has 31 char limit
                        
                        # Write each DataFrame with appropriate styling
                        start_row = 0
                        for df, df_type in all_dfs:
                            df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)
                            
                            # Get the worksheet
                            worksheet = writer.sheets[sheet_name]
                            
                            # Apply styles based on the type of content
                            if df_type == 'header':
                                # Style for main headers
                                for cell in worksheet[start_row + 1]:
                                    cell.font = openpyxl.styles.Font(bold=True, size=14)
                                    cell.fill = openpyxl.styles.PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                            
                            elif df_type == 'fact_title':
                                # Style for fact titles
                                for cell in worksheet[start_row + 1]:
                                    cell.font = openpyxl.styles.Font(bold=True, size=12)
                                    cell.fill = openpyxl.styles.PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                            
                            elif df_type == 'fact_data':
                                # Style for fact data headers
                                for cell in worksheet[start_row + 1]:
                                    cell.font = openpyxl.styles.Font(bold=True)
                                    cell.fill = openpyxl.styles.PatternFill(start_color='F8F8F8', end_color='F8F8F8', fill_type='solid')
                                
                                # Add borders to the data table
                                table_rows = worksheet[start_row + 1:start_row + len(df) + 1]
                                thin_border = openpyxl.styles.Border(
                                    left=openpyxl.styles.Side(style='thin'),
                                    right=openpyxl.styles.Side(style='thin'),
                                    top=openpyxl.styles.Side(style='thin'),
                                    bottom=openpyxl.styles.Side(style='thin')
                                )
                                for row in table_rows:
                                    for cell in row:
                                        cell.border = thin_border
                            
                            elif df_type == 'details':
                                # Style for insight details
                                for row in worksheet[start_row + 1:start_row + len(df) + 1]:
                                    for cell in row:
                                        cell.font = openpyxl.styles.Font(size=11)
                                        if cell.column == 1:  # Property column
                                            cell.font = openpyxl.styles.Font(bold=True)
                            
                            # Adjust column widths
                            for column in worksheet.columns:
                                max_length = 0
                                column = [cell for cell in column]
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(cell.value)
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                worksheet.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)
                            
                            start_row += len(df) + 2  # Add 2 for spacing
                
                # If no sheets were created, add a No Data sheet
                if not use_case_insights:
                    df = pd.DataFrame({'Message': ['No data available']})
                    df.to_excel(writer, sheet_name='No Data', index=False)
        
        except Exception as e:
            print(f"Error while writing Excel file: {str(e)}")
            return jsonify({'error': f'Error creating Excel file: {str(e)}'}), 500
        
        # Seek to the beginning of the file
        output.seek(0)
        
        try:
            response = send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            response.headers['Cache-Control'] = 'no-cache'
            return response
            
        except Exception as e:
            print(f"Error sending file: {str(e)}")
            return jsonify({'error': f'Error sending file: {str(e)}'}), 500
        
    except Exception as e:
        print(f"Error in download_excel: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Development only
    app.debug = True
    port = int(os.environ.get('PORT', 10001))
    app.run(host='0.0.0.0', port=port)
